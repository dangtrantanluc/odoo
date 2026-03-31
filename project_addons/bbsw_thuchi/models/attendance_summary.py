import calendar
from datetime import date, timedelta
from odoo import models, fields, api
from odoo.exceptions import UserError


class BbswAttendanceSummaryWizard(models.TransientModel):
    _name = 'bbsw.attendance.summary.wizard'
    _description = 'Tổng hợp chấm công tháng'

    period_month = fields.Integer(
        string='Tháng', required=True,
        default=lambda self: fields.Date.today().month)
    period_year = fields.Integer(
        string='Năm', required=True,
        default=lambda self: fields.Date.today().year)
    department_id = fields.Many2one(
        'hr.department', string='Phòng ban',
        help='Để trống = tất cả phòng ban')
    use_payroll_period = fields.Boolean(
        string='Theo kỳ lương (26→27)',
        default=True,
        help='Tích: kỳ 26 tháng trước → 27 tháng này. Bỏ tích: ngày 1 → cuối tháng.')

    def _get_date_range(self):
        """Trả về (date_from, date_to) dạng date object."""
        m, y = self.period_month, self.period_year
        if self.use_payroll_period:
            if m == 1:
                date_from = date(y - 1, 12, 26)
            else:
                date_from = date(y, m - 1, 26)
            date_to = date(y, m, 27)
        else:
            date_from = date(y, m, 1)
            last_day = calendar.monthrange(y, m)[1]
            date_to = date(y, m, last_day)
        return date_from, date_to

    def _get_employees(self):
        domain = [('work_status', '=', 'WORKING'), ('active', '=', True)]
        if self.department_id:
            domain.append(('department_id', '=', self.department_id.id))
        return self.env['hr.employee'].search(domain, order='department_id, name')

    def action_export_excel(self):
        """Xuất file Excel bảng chấm công tháng."""
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError('Cần cài openpyxl để xuất Excel.')

        import base64
        from io import BytesIO

        m, y = self.period_month, self.period_year
        date_from, date_to = self._get_date_range()
        employees = self._get_employees()

        if not employees:
            raise UserError('Không tìm thấy nhân viên nào.')

        # Danh sách ngày trong kỳ
        days = []
        d = date_from
        while d <= date_to:
            days.append(d)
            d += timedelta(days=1)

        # Đọc toàn bộ attendance trong kỳ
        att_domain = [
            ('check_in', '>=', str(date_from)),
            ('check_in', '<=', str(date_to) + ' 23:59:59'),
        ]
        if self.department_id:
            att_domain.append(('employee_id.department_id', '=', self.department_id.id))
        attendances = self.env['hr.attendance'].sudo().search(att_domain)

        # Index: {(employee_id, date): [attendance_records]}
        att_index = {}
        for att in attendances:
            emp_id = att.employee_id.id
            att_date = att.check_in.date()
            key = (emp_id, att_date)
            if key not in att_index:
                att_index[key] = []
            att_index[key].append(att)

        # ── Tạo workbook ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Chấm công T{m:02d}-{y}'

        # ── Styles ──
        def make_fill(hex_color):
            return PatternFill('solid', fgColor=hex_color)

        def thin_border():
            s = Side(style='thin', color='BFBFBF')
            return Border(left=s, right=s, top=s, bottom=s)

        title_font   = Font(name='Arial', bold=True, size=13, color='1F3864')
        header_font  = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        sub_font     = Font(name='Arial', bold=True, size=9, color='1F3864')
        data_font    = Font(name='Arial', size=9)
        total_font   = Font(name='Arial', bold=True, size=9)
        center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left         = Alignment(horizontal='left', vertical='center')

        header_fill  = make_fill('1F3864')
        weekend_fill = make_fill('F2F2F2')
        holiday_fill = make_fill('FCE4D6')
        normal_fill  = make_fill('FFFFFF')
        total_fill   = make_fill('DDEBF7')
        title_fill   = make_fill('DEEAF1')

        WEEKEND_COLOR = 'BFBFBF'
        CODE_COLORS = {
            'X':   ('E2EFDA', '375623'),  # xanh nhạt
            'P':   ('FFF2CC', '7F6000'),  # vàng
            'NL':  ('FCE4D6', '833C00'),  # cam nhạt
            'KL':  ('FFE7E7', 'C00000'),  # đỏ nhạt
            'TV':  ('EAF0FB', '1F497D'),  # xanh dương nhạt
            'BH':  ('E7F3E7', '375623'),
        }

        # ── Row 1: Tiêu đề ──
        title_text = f'BẢNG CHẤM CÔNG THÁNG {m:02d}/{y}'
        period_text = f'Kỳ lương: {date_from.strftime("%d/%m/%Y")} → {date_to.strftime("%d/%m/%Y")}'

        # Fixed columns: STT, Họ tên, Mã NV, Phòng ban
        FIXED_COLS = 4
        total_cols = FIXED_COLS + len(days) + 6  # +6 summary cols

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(1, 1, title_text)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        period_cell = ws.cell(2, 1, period_text)
        period_cell.font = sub_font
        period_cell.fill = title_fill
        period_cell.alignment = center

        # ── Row 3: Header tháng/năm + header cột ──
        ws.row_dimensions[3].height = 30
        headers_fixed = ['STT', 'Họ tên', 'Mã NV', 'Phòng ban']
        for ci, h in enumerate(headers_fixed, 1):
            c = ws.cell(3, ci, h)
            c.font = header_font
            c.fill = make_fill('1F3864')
            c.alignment = center
            c.border = thin_border()

        # Day columns
        WEEKDAY_VN = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
        for di, day in enumerate(days):
            col = FIXED_COLS + di + 1
            is_weekend = day.weekday() >= 5
            label = f'{day.day}\n{WEEKDAY_VN[day.weekday()]}'
            c = ws.cell(3, col, label)
            c.font = Font(name='Arial', bold=True, size=8,
                          color='FFFFFF' if not is_weekend else '666666')
            c.fill = make_fill('1F3864') if not is_weekend else make_fill('4F5B62')
            c.alignment = center
            c.border = thin_border()
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Summary columns
        summary_headers = ['Ngày X', 'Ngày P', 'Ngày NL', 'Ngày KL', 'Tổng giờ', 'Ghi chú']
        summary_colors  = ['375623', '7F6000', '833C00', 'C00000', '1F497D', '404040']
        for si, (sh, sc) in enumerate(zip(summary_headers, summary_colors)):
            col = FIXED_COLS + len(days) + si + 1
            c = ws.cell(3, col, sh)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = make_fill('2E4057')
            c.alignment = center
            c.border = thin_border()
            ws.column_dimensions[get_column_letter(col)].width = 9

        # ── Data rows ──
        for row_idx, emp in enumerate(employees, 1):
            row = 3 + row_idx
            ws.row_dimensions[row].height = 18

            # Fixed cols
            for ci, val in enumerate([row_idx, emp.name,
                                       emp.employee_code or '',
                                       emp.department_id.name or ''], 1):
                c = ws.cell(row, ci, val)
                c.font = data_font if ci > 1 else Font(name='Arial', size=9, bold=True)
                c.alignment = center if ci != 2 else left
                c.border = thin_border()

            # Day cols
            total_x = total_p = total_nl = total_kl = total_h = 0.0

            for di, day in enumerate(days):
                col = FIXED_COLS + di + 1
                is_weekend = day.weekday() >= 5
                key = (emp.id, day)
                records = att_index.get(key, [])

                if is_weekend:
                    cell_val = 'CN' if day.weekday() == 6 else 'T7'
                    bg = 'F2F2F2'
                    fg = '999999'
                elif not records:
                    # Không có record trong ngày làm việc
                    cell_val = ''
                    bg = 'FFFFFF'
                    fg = '000000'
                else:
                    # Lấy mã chủ đạo trong ngày (ưu tiên mã đầu tiên có code)
                    codes = [r.attendance_code or 'X' for r in records]
                    hours = sum(r.worked_hours for r in records)
                    total_h += hours
                    main_code = codes[0].upper() if codes else 'X'

                    # Đếm ngày
                    if main_code == 'X' or main_code not in ('P', 'NL', 'KL', 'TV', 'BH', 'NB'):
                        total_x += 1
                    elif main_code == 'P':
                        total_p += 1
                    elif main_code == 'NL':
                        total_nl += 1
                    elif main_code == 'KL':
                        total_kl += 1

                    cell_val = main_code
                    code_style = CODE_COLORS.get(main_code, ('FFFFFF', '000000'))
                    bg, fg = code_style

                c = ws.cell(row, col, cell_val)
                c.font = Font(name='Arial', size=8, bold=bool(records and not is_weekend),
                              color=fg)
                c.fill = make_fill(bg)
                c.alignment = center
                c.border = thin_border()

            # Summary cols
            summary_vals = [total_x, total_p, total_nl, total_kl,
                            round(total_h, 1), '']
            for si, sv in enumerate(summary_vals):
                col = FIXED_COLS + len(days) + si + 1
                c = ws.cell(row, col, sv)
                c.font = total_font
                c.fill = total_fill
                c.alignment = center
                c.border = thin_border()

        # ── Totals row ──
        total_row = 3 + len(employees) + 1
        ws.row_dimensions[total_row].height = 20
        ws.merge_cells(start_row=total_row, start_column=1,
                       end_row=total_row, end_column=FIXED_COLS)
        tc = ws.cell(total_row, 1, 'TỔNG CỘNG')
        tc.font = Font(name='Arial', bold=True, size=9, color='1F3864')
        tc.fill = make_fill('BDD7EE')
        tc.alignment = center
        tc.border = thin_border()

        # Fixed col widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 16
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 16

        ws.freeze_panes = 'E4'  # Freeze first 4 cols + header rows

        # ── Legend sheet ──
        ws2 = wb.create_sheet('Chú thích')
        legends = [
            ('Mã', 'Ý nghĩa', 'Màu'),
            ('X', 'Ngày công chính thức', '#E2EFDA'),
            ('P', 'Nghỉ phép có lương', '#FFF2CC'),
            ('NL', 'Nghỉ lễ', '#FCE4D6'),
            ('KL', 'Nghỉ không lương', '#FFE7E7'),
            ('TV', 'Thử việc', '#EAF0FB'),
            ('BH', 'Bảo hiểm', '#E7F3E7'),
            ('T7/CN', 'Cuối tuần', '#F2F2F2'),
            ('(trống)', 'Chưa có dữ liệu chấm công', '#FFFFFF'),
        ]
        for ri, (a, b, _) in enumerate(legends, 1):
            ws2.cell(ri, 1, a).font = Font(bold=(ri == 1))
            ws2.cell(ri, 2, b).font = Font(bold=(ri == 1))
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 35

        # ── Save ──
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = base64.b64encode(buf.read())

        filename = f'Cham_cong_T{m:02d}_{y}.xlsx'

        # Save as attachment
        attach = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attach.id}?download=true',
            'target': 'self',
        }

    def action_view_summary(self):
        """Tổng hợp chấm công theo nhân viên và mở bảng tổng hợp."""
        import pytz
        from collections import defaultdict
        date_from, date_to = self._get_date_range()
        employees = self._get_employees()

        att_domain = [
            ('check_in', '>=', str(date_from)),
            ('check_in', '<=', str(date_to) + ' 23:59:59'),
        ]
        if self.department_id:
            att_domain.append(('employee_id.department_id', '=', self.department_id.id))
        attendances = self.env['hr.attendance'].sudo().search(att_domain)

        # Đếm ngày làm việc trong kỳ (T2→T6)
        working_days_in_period = sum(
            1 for i in range((date_to - date_from).days + 1)
            if (date_from + timedelta(days=i)).weekday() < 5
        )

        # Nhóm chấm công theo nhân viên
        emp_atts = defaultdict(list)
        for att in attendances:
            emp_atts[att.employee_id.id].append(att)

        lines = []
        for emp in employees:
            atts = emp_atts.get(emp.id, [])
            days_x = days_p = days_nl = days_kl = days_tv = 0
            total_hours = 0.0
            seen_dates = set()

            for att in atts:
                att_date = att.check_in.date()
                code = (att.attendance_code or 'X').upper()
                total_hours += att.worked_hours or 0
                if att_date not in seen_dates:
                    seen_dates.add(att_date)
                    if code not in ('P', 'NL', 'KL', 'TV', 'BH', 'NB'):
                        days_x += 1
                    elif code == 'P':
                        days_p += 1
                    elif code == 'NL':
                        days_nl += 1
                    elif code == 'KL':
                        days_kl += 1
                    elif code == 'TV':
                        days_tv += 1

            days_missing = max(0, working_days_in_period - len(seen_dates))

            lines.append({
                'wizard_id': self.id,
                'employee_id': emp.id,
                'employee_name': emp.name,
                'employee_code': emp.employee_code or '',
                'department_name': emp.department_id.name or '',
                'period_label': f'T{self.period_month:02d}/{self.period_year}',
                'days_x': days_x,
                'days_p': days_p,
                'days_nl': days_nl,
                'days_kl': days_kl,
                'days_tv': days_tv,
                'total_hours': round(total_hours, 1),
                'checkin_count': len(atts),
                'days_missing': days_missing,
            })

        created_lines = self.env['bbsw.attendance.summary.line'].sudo().create(lines) if lines else self.env['bbsw.attendance.summary.line']

        # Tạo day lines cho từng nhân viên
        WEEKDAY_VN = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
        vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')

        def to_vn_time(dt):
            if not dt:
                return ''
            return dt.replace(tzinfo=pytz.utc).astimezone(vn_tz).strftime('%H:%M')

        day_line_data = []
        for summary_line in created_lines:
            emp_id = summary_line.employee_id.id
            emp_atts_by_date = defaultdict(list)
            for att in emp_atts.get(emp_id, []):
                emp_atts_by_date[att.check_in.date()].append(att)

            d = date_from
            while d <= date_to:
                is_weekend = d.weekday() >= 5
                records = emp_atts_by_date.get(d, [])

                if records:
                    main_att = sorted(records, key=lambda r: r.check_in)[0]
                    code = (main_att.attendance_code or 'X').upper()
                    check_in_str = to_vn_time(main_att.check_in)
                    check_out_str = to_vn_time(main_att.check_out) if main_att.check_out else ''
                    hours = round(sum(r.worked_hours for r in records), 1)
                    att_id = main_att.id
                else:
                    code = 'CN' if d.weekday() == 6 else ('T7' if d.weekday() == 5 else '')
                    check_in_str = check_out_str = ''
                    hours = 0.0
                    att_id = False

                day_line_data.append({
                    'summary_id': summary_line.id,
                    'attendance_id': att_id or False,
                    'date': d,
                    'weekday_label': WEEKDAY_VN[d.weekday()],
                    'is_weekend': is_weekend,
                    'attendance_code': code,
                    'check_in_time': check_in_str,
                    'check_out_time': check_out_str,
                    'worked_hours': hours,
                })
                d += timedelta(days=1)

        if day_line_data:
            self.env['bbsw.attendance.day.line'].sudo().create(day_line_data)

        return {
            'type': 'ir.actions.act_window',
            'name': f'Tổng hợp chấm công T{self.period_month:02d}/{self.period_year}',
            'res_model': 'bbsw.attendance.summary.line',
            'view_mode': 'tree,form',
            'domain': [('wizard_id', '=', self.id)],
            'target': 'current',
        }


class BbswAttendanceSummaryLine(models.TransientModel):
    _name = 'bbsw.attendance.summary.line'
    _description = 'Dòng tổng hợp chấm công'
    _order = 'department_name, employee_name'

    wizard_id = fields.Many2one('bbsw.attendance.summary.wizard', string='Wizard')
    employee_id = fields.Many2one('hr.employee', string='Nhân viên')
    employee_name = fields.Char(string='Họ tên')
    employee_code = fields.Char(string='Mã NV')
    department_name = fields.Char(string='Phòng ban')
    period_label = fields.Char(string='Kỳ')
    days_x = fields.Integer(string='Ngày X', help='Số ngày công chính thức')
    days_p = fields.Integer(string='Ngày P', help='Số ngày nghỉ phép có lương')
    days_nl = fields.Integer(string='Ngày NL', help='Số ngày nghỉ lễ')
    days_kl = fields.Integer(string='Ngày KL', help='Số ngày nghỉ không lương')
    days_tv = fields.Integer(string='Ngày TV', help='Số ngày thử việc')
    total_hours = fields.Float(string='Tổng giờ', digits=(6, 1))
    checkin_count = fields.Integer(string='Lần chấm')
    days_missing = fields.Integer(string='Thiếu dữ liệu',
                                  help='Ngày làm việc chưa có bản ghi chấm công')
    day_ids = fields.One2many(
        'bbsw.attendance.day.line', 'summary_id', string='Chi tiết từng ngày')

    def action_save_attendance_codes(self):
        """Ghi mã chấm công đã sửa về hr.attendance."""
        from datetime import datetime, time as dtime
        for day in self.day_ids:
            if day.is_weekend:
                continue
            code = (day.attendance_code or '').upper().strip()
            if not code:
                continue
            if day.attendance_id:
                day.attendance_id.sudo().write({'attendance_code': code})
            elif code not in ('X', ''):
                # Tạo record mới cho ngày vắng (P, NL, KL...)
                check_in_dt = datetime.combine(day.date, dtime(2, 0))
                self.env['hr.attendance'].sudo().create({
                    'employee_id': self.employee_id.id,
                    'check_in': check_in_dt,
                    'check_out': check_in_dt,
                    'attendance_code': code,
                })
        return {'type': 'ir.actions.client', 'tag': 'reload'}

    def action_back_to_list(self):
        """Quay lại danh sách tổng hợp."""
        return {
            'type': 'ir.actions.act_window',
            'name': f'Tổng hợp chấm công {self.period_label}',
            'res_model': 'bbsw.attendance.summary.line',
            'view_mode': 'tree,form',
            'domain': [('wizard_id', '=', self.wizard_id.id)],
            'target': 'current',
        }


class BbswAttendanceDayLine(models.TransientModel):
    _name = 'bbsw.attendance.day.line'
    _description = 'Chi tiết chấm công từng ngày'
    _order = 'date asc'

    summary_id = fields.Many2one(
        'bbsw.attendance.summary.line', string='Tổng hợp', ondelete='cascade')
    attendance_id = fields.Many2one('hr.attendance', string='Bản ghi CC')
    date = fields.Date(string='Ngày')
    weekday_label = fields.Char(string='Thứ')
    is_weekend = fields.Boolean(string='Cuối tuần')
    attendance_code = fields.Char(string='Mã CC', help='X=Công | P=Phép | NL=Nghỉ lễ | KL=Không lương | TV=Thử việc')
    check_in_time = fields.Char(string='Giờ vào')
    check_out_time = fields.Char(string='Giờ ra')
    worked_hours = fields.Float(string='Số giờ', digits=(4, 1))
    note = fields.Char(string='Ghi chú')
